"""
Excel parser. Streams `.xlsx` bytes through openpyxl in `read_only=True`
mode so the workbook is never fully materialized in memory, and in
`data_only=True` mode so formulas are read as their cached values and no
formula re-evaluation occurs (also a defence against formula-injection
patterns saved in malicious files).

The parser does **not** validate values — it only checks the header row
and produces one `RawRow` per data row with whatever openpyxl returned for
each cell. Type-checking and value-checking happen in `ingestion.validator`.
"""

from __future__ import annotations

import io
from collections.abc import Iterator
from contextlib import contextmanager
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from src.domain.models import RawRow

EXPECTED_COLUMNS: frozenset[str] = frozenset(
    {"ClientId", "TransactionId", "ISIN", "Action", "Quantity", "Price", "Timestamp"}
)


class HeaderValidationError(ValueError):
    """Raised when the workbook can't be opened or its header row is malformed.

    Distinct from row-level validation errors — a bad header means no rows
    can be processed, so the API layer returns 422 with a single message
    rather than a per-row error list.
    """


def parse_workbook(content: bytes) -> list[RawRow]:
    """
    Parse the bytes of an `.xlsx` file into a list of `RawRow`.

    Raises:
        HeaderValidationError: file isn't a readable xlsx, has no active
            sheet, is empty, or is missing one or more expected columns.
    """
    with _open_workbook(content) as workbook:
        sheet = workbook.active
        if sheet is None:
            raise HeaderValidationError("Workbook has no active sheet")

        rows_iter = sheet.iter_rows(values_only=True)
        col_index = _read_header_and_build_index(rows_iter)
        return _parse_data_rows(rows_iter, col_index)


@contextmanager
def _open_workbook(content: bytes) -> Iterator[Workbook]:
    """Load *content* as a read-only openpyxl workbook and ensure it closes.

    Deliberately swallows the underlying openpyxl exception text in the
    user-facing message — openpyxl errors can include implementation details
    or file-system paths that we don't want surfaced through the API. The
    original exception is preserved as the `__cause__` of the
    `HeaderValidationError` for server-side debugging via `raise … from exc`.
    """
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HeaderValidationError(
            "Could not read workbook — the file may be corrupt or not a valid .xlsx"
        ) from exc

    try:
        yield workbook
    finally:
        workbook.close()


def _read_header_and_build_index(
    rows_iter: Iterator[tuple[Any, ...]],
) -> dict[str, int]:
    """
    Consume the header row from *rows_iter* and return a column-name → index map.

    Raises `HeaderValidationError` if the workbook is empty, the header row
    contains only blanks, any expected column is missing, or any expected
    column appears more than once. Column order in the spreadsheet does not
    matter — the returned map lets downstream code look cells up by name.
    """
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise HeaderValidationError("Workbook is empty (no header row)") from exc

    if header_row is None or all(cell is None for cell in header_row):
        raise HeaderValidationError("Header row is empty")

    header_values = [str(cell).strip() if cell is not None else "" for cell in header_row]
    missing = EXPECTED_COLUMNS - set(header_values)
    if missing:
        raise HeaderValidationError(f"Missing required columns: {sorted(missing)}")

    # Detect duplicate headers explicitly — `list.index()` would silently
    # return the first occurrence and the second copy would be ignored.
    col_index: dict[str, int] = {}
    duplicates: list[str] = []
    for idx, name in enumerate(header_values):
        if name not in EXPECTED_COLUMNS:
            continue
        if name in col_index:
            duplicates.append(name)
        else:
            col_index[name] = idx
    if duplicates:
        raise HeaderValidationError(f"Duplicate column headers: {sorted(set(duplicates))}")

    return col_index


def _parse_data_rows(
    rows_iter: Iterator[tuple[Any, ...]],
    col_index: dict[str, int],
) -> list[RawRow]:
    """
    Build one `RawRow` per non-empty data row. Spreadsheet rows are 1-indexed,
    and the header is row 1, so the first data row is row number 2.
    """
    rows: list[RawRow] = []
    for row_number, row_values in enumerate(rows_iter, start=2):
        if _is_blank_row(row_values):
            continue
        rows.append(_build_raw_row(row_number, row_values, col_index))
    return rows


def _is_blank_row(row_values: tuple[Any, ...] | None) -> bool:
    """True for trailing rows that openpyxl yields as all-None tuples."""
    return row_values is None or all(cell is None for cell in row_values)


def _build_raw_row(
    row_number: int,
    row_values: tuple[Any, ...],
    col_index: dict[str, int],
) -> RawRow:
    """Pull each expected column from *row_values* into a `RawRow`."""
    return RawRow(
        row_number=row_number,
        transaction_id=_cell(row_values, col_index["TransactionId"]),
        client_id=_cell(row_values, col_index["ClientId"]),
        isin=_cell(row_values, col_index["ISIN"]),
        action=_cell(row_values, col_index["Action"]),
        quantity=_cell(row_values, col_index["Quantity"]),
        price=_cell(row_values, col_index["Price"]),
        timestamp=_cell(row_values, col_index["Timestamp"]),
    )


def _cell(values: tuple[Any, ...], idx: int) -> Any:
    """Return the cell at *idx*, or None if the row is shorter than that index."""
    if idx >= len(values):
        return None
    return values[idx]
